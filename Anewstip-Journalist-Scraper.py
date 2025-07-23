import tkinter as tk
from tkinter import filedialog, messagebox
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import NoSuchElementException
from bs4 import BeautifulSoup
import pandas as pd
import time
import threading

driver = None
save_path = None
keywords = []

def load_keywords():
    global keywords
    file_path = filedialog.askopenfilename(
        title="Select Keywords .txt File",
        filetypes=[("Text files", "*.txt")]
    )
    if not file_path:
        messagebox.showerror("Error", "No keyword file selected.")
        return
    with open(file_path, 'r', encoding='utf-8') as f:
        keywords = [line.strip() for line in f if line.strip()]
    keyword_label.config(text=f"✔ Loaded {len(keywords)} keywords")

def choose_file():
    global save_path
    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")],
        title="Save Extracted Data As"
    )
    if file_path:
        save_path = file_path
        file_label.config(text=f"✔ File: {save_path}")
    else:
        file_label.config(text="❌ No file selected")

def launch_browser():
    global driver
    if not keywords:
        messagebox.showerror("Missing Keywords", "Please load a keywords .txt file first.")
        return

    options = Options()
    options.add_experimental_option("detach", True)
    options.add_argument("--start-maximized")
    driver = webdriver.Chrome(service=Service(), options=options)
    messagebox.showinfo("Login Required", "Browser launched.\nLogin to Anewstip manually.\nThen click 'Start Extract'.")

def start_scraping():
    global driver, save_path, keywords
    if not driver:
        messagebox.showerror("Error", "Browser not launched. Click 'Launch Anewstip' first.")
        return
    if not save_path:
        messagebox.showerror("Error", "Please choose a file to save the data.")
        return
    if not keywords:
        messagebox.showerror("Error", "Please load a keywords .txt file.")
        return

    def scrape_all():
        headers = ['Keyword', 'Name', 'Title', 'Outlet', 'Email', 'Phone', 'Topics']
        df = pd.DataFrame(columns=headers)
        df.to_excel(save_path, index=False)

        from openpyxl import load_workbook
        wb = load_workbook(save_path)
        ws = wb.active

        # Load previously saved rows to skip duplicates
        existing_rows = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            existing_rows.add(tuple(row))

        for kw in keywords:
            print(f"\n🔍 Searching keyword: {kw}")
            search_url = f"https://anewstip.com/search/journalists/?q={kw}"
            driver.get(search_url)
            time.sleep(4)

            visited_pages = set()

            while True:
                soup = BeautifulSoup(driver.page_source, 'html.parser')

                # Handle empty result
                no_result_msg = soup.find(string="Your search didn't match any results.")
                if no_result_msg:
                    print(f"⚠️ No results found for keyword: {kw}")
                    break

                current_page = '1'
                current_page_tag = soup.select_one('span.page-link.current-page')
                if current_page_tag:
                    current_page = current_page_tag.text.strip()
                    visited_pages.add(current_page)

                profiles = soup.select('.user-profile')
                if not profiles:
                    print(f"⚠️ No profiles found on page {current_page} for keyword: {kw}")
                    break

                for profile in profiles:
                    try: name = profile.select_one('.info-name a').get_text(strip=True)
                    except: name = ''
                    try: title = profile.select_one('.info-title a').get_text(strip=True)
                    except: title = ''
                    try: outlet = profile.select_one('.info-outlet-name a').get_text(strip=True)
                    except: outlet = ''
                    try: email = profile.select_one('.info-email a').get_text(strip=True)
                    except: email = ''
                    try: phone = profile.select_one('.info-phone span').get_text(strip=True)
                    except: phone = ''
                    try:
                        topics = ', '.join(
                            tag.get_text(strip=True)
                            for tag in profile.select('.topic-list .topic-label')
                        )
                    except: topics = ''

                    new_row = (kw, name, title, outlet, email, phone, topics)
                    if new_row not in existing_rows:
                        ws.append(new_row)
                        wb.save(save_path)
                        existing_rows.add(new_row)
                    else:
                        print(f"⏩ Skipped duplicate: {name} ({email})")

                # Try clicking next unvisited page
                next_page = None
                try:
                    all_page_links = driver.find_elements(By.CSS_SELECTOR, 'a.page-link[data-page]')
                    for link in all_page_links:
                        page_num = link.get_attribute('data-page')
                        if page_num and page_num not in visited_pages:
                            next_page = link
                            break
                except:
                    break

                if next_page:
                    try:
                        next_page.click()
                        time.sleep(4)
                    except:
                        break
                else:
                    break  # All pages visited

        wb.save(save_path)
        messagebox.showinfo("Done", f"✅ Scraping complete.\nSaved to:\n{save_path}")

    threading.Thread(target=scrape_all).start()

# GUI setup
root = tk.Tk()
root.title("Anewstip Journalist Scraper")
root.geometry("450x400")

tk.Label(root, text="🗂 Load .txt file with keywords (one per line):").pack(pady=(10, 0))
tk.Button(root, text="📂 Load Keywords File", command=load_keywords, bg="#FFC107", fg="black").pack(pady=5)
keyword_label = tk.Label(root, text="❌ No keyword file loaded")
keyword_label.pack()

tk.Button(root, text="Choose Save Location", command=choose_file, bg="#673AB7", fg="white").pack(pady=10)
file_label = tk.Label(root, text="❌ No file selected")
file_label.pack()

tk.Button(root, text="Launch Anewstip", command=launch_browser, bg="#4CAF50", fg="white", height=2, width=30).pack(pady=10)
tk.Button(root, text="Start Extract", command=start_scraping, bg="#2196F3", fg="white", height=2, width=30).pack(pady=5)

tk.Label(root, text="ℹ️ Excel updates live during scraping.\nPlease login manually before scraping.").pack(pady=20)

root.mainloop()
